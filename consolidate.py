"""Reads every JSON extraction in output/json/ and writes, per report set
(core = every report; one more per detected format):

  - output/all_reports_core.csv / .xlsx
  - output/all_reports_<format>.csv / .xlsx -- same core columns plus that
    format's own extra fields (safe_number, dbt, ... for Creditsafe;
    points_allocated, rating_band, ... for CRIF; gladtrust_rating,
    unified_social_credit_code, ... for GLADTRUST).

Full detail -- every director, every year of financials, every adverse
finding -- stays in the per-report JSON; these are for a quick sort/filter
pass across the whole batch, not the source of truth.

Both formats carry the same rows; they exist for different failure modes:

* CSV is written with utf-8-sig (a UTF-8 BOM) so Excel renders non-Latin
  text (Chinese, Arabic, ...) instead of mojibake -- but CSV has no way to
  mark a column's type, so Excel still auto-detects one from the cell text
  on open: a value like "1-9" (employees) becomes a date, and an 18-digit
  registration/USCC number becomes scientific notation -- and if the file is
  then saved, that precision loss is permanent (Excel's numeric precision
  caps at 15 significant digits).
* XLSX carries real per-cell format metadata, so every text column is
  written with an explicit Text format and Excel never reinterprets it.
  Open the .xlsx, not the .csv, when working in Excel.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import config

_FORMAT_EXTRA_FIELDNAMES = {
    config.FORMAT_CREDITSAFE: [
        "safe_number",
        "probability_of_default",
        "contract_limit",
        "dbt",
        "enquiries_past_12_months",
        "share_capital",
    ],
    config.FORMAT_CRIF_INTL: [
        "points_allocated",
        "rating_band",
        "payment_terms",
        "corruption_index_summary",
    ],
    config.FORMAT_GLADTRUST: [
        "gladtrust_rating",
        "gladtrust_suggestion",
        "base_credit_limit",
        "unified_social_credit_code",
        "business_scope",
    ],
}

_MONEY_FIELDNAMES = {"contract_limit", "share_capital", "base_credit_limit"}
_NUMERIC_FIELDNAMES = {"directors_count", "shareholders_count", "adverse_findings_count"}

_CORE_FIELDNAMES = [
    "source_file",
    "format",
    "company_name",
    "name_local",
    "country",
    "registration_number",
    "tax_vat_number",
    "incorporation_date",
    "legal_form",
    "company_status",
    "registered_address",
    "industry_description",
    "employees",
    "report_date",
    "credit_score",
    "score_scale",
    "risk_level",
    "credit_rating",
    "credit_limit",
    "directors_count",
    "shareholders_count",
    "adverse_findings_count",
    "latest_financial_year",
    "latest_turnover",
    "latest_pre_tax_profit",
    "latest_net_worth",
]


def _money(value: dict | None) -> str:
    if not value:
        return ""
    return value.get("raw") or ("" if value.get("value") is None else str(value["value"]))


def _credit_limit(report: dict) -> str:
    """credit_limit should always be populated per schemas.py's field
    description, but falls back to the format-specific credit-limit field
    (base_credit_limit for GLADTRUST, contract_limit for Creditsafe) in case
    an older extraction predates that fix."""
    for field in ("credit_limit", "base_credit_limit", "contract_limit"):
        value = _money(report.get(field))
        if value:
            return value
    return ""


def _latest_financial_year(financials: list[dict]) -> dict:
    years = [f for f in financials if f.get("year")]
    if years:
        return max(years, key=lambda f: f["year"])
    return financials[0] if financials else {}


def _core_row(report: dict) -> dict:
    latest = _latest_financial_year(report.get("financials") or [])
    return {
        "source_file": report.get("source_file", ""),
        "format": report.get("format", ""),
        "company_name": report.get("company_name") or "",
        "name_local": report.get("name_local") or "",
        "country": report.get("country") or "",
        "registration_number": report.get("registration_number") or "",
        "tax_vat_number": report.get("tax_vat_number") or "",
        "incorporation_date": report.get("incorporation_date") or "",
        "legal_form": report.get("legal_form") or "",
        "company_status": report.get("company_status") or "",
        "registered_address": report.get("registered_address") or "",
        "industry_description": report.get("industry_description") or "",
        "employees": report.get("employees") or "",
        "report_date": report.get("report_date") or "",
        "credit_score": report.get("credit_score") or "",
        "score_scale": report.get("score_scale") or "",
        "risk_level": report.get("risk_level") or "",
        "credit_rating": report.get("credit_rating") or "",
        "credit_limit": _credit_limit(report),
        "directors_count": len(report.get("directors") or []),
        "shareholders_count": len(report.get("shareholders") or []),
        "adverse_findings_count": len(report.get("adverse_findings") or []),
        "latest_financial_year": latest.get("year", ""),
        "latest_turnover": _money(latest.get("turnover")),
        "latest_pre_tax_profit": _money(latest.get("pre_tax_profit")),
        "latest_net_worth": _money(latest.get("net_worth")),
    }


def _extra_row(report: dict, extra_fields: list[str]) -> dict:
    return {
        field: (_money(report.get(field)) if field in _MONEY_FIELDNAMES else report.get(field) or "")
        for field in extra_fields
    }


def _write_csv(path: Path, fieldnames: list[str], rows: list[dict], log=print) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    log(f"{len(rows)} reports -> {path}")


def _write_xlsx(path: Path, fieldnames: list[str], rows: list[dict], log=print) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(name, "") for name in fieldnames])

    text_columns = (i for i, name in enumerate(fieldnames, start=1) if name not in _NUMERIC_FIELDNAMES)
    for col_idx in text_columns:
        for cell in ws[get_column_letter(col_idx)][1:]:  # [1:] skips the header row
            cell.number_format = "@"

    wb.save(path)
    log(f"{len(rows)} reports -> {path}")


def _write(base_path: Path, fieldnames: list[str], rows: list[dict], log=print) -> None:
    _write_csv(base_path.with_suffix(".csv"), fieldnames, rows, log)
    _write_xlsx(base_path.with_suffix(".xlsx"), fieldnames, rows, log)


def consolidate(json_dir: Path, core_csv_path: Path, *, log=print) -> int:
    """Writes the CSV/XLSX summaries. Returns the number of reports included."""
    reports = [json.loads(p.read_text(encoding="utf-8")) for p in sorted(json_dir.glob("*.json"))]

    _write(core_csv_path, _CORE_FIELDNAMES, [_core_row(r) for r in reports], log)

    output_dir = core_csv_path.parent
    for fmt, extra_fields in _FORMAT_EXTRA_FIELDNAMES.items():
        fmt_reports = [r for r in reports if r.get("format") == fmt]
        if not fmt_reports:
            continue
        rows = [{**_core_row(r), **_extra_row(r, extra_fields)} for r in fmt_reports]
        _write(output_dir / f"all_reports_{fmt}.csv", _CORE_FIELDNAMES + extra_fields, rows, log)

    return len(reports)


if __name__ == "__main__":
    output_dir = Path(__file__).parent / "output"
    consolidate(output_dir / config.OUTPUT_JSON_DIRNAME, output_dir / config.CORE_CSV_FILENAME)
